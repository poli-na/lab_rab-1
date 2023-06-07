from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook
import os


def parse1():
    filename = 'лабалоторная1.xlsx'
    exelfile = load_workbook(filename)
    spisok = exelfile['данные']
    description1 = []
    description2 = []
    description3 = []
    count = 2
    ssilka = 'https://www.chitai-gorod.ru/catalog/collections/bestsell?page='

    for i in range(1, count + 1):
        url1 = ssilka + str(i)
        page1 = requests.get(url1)
        print(page1.status_code)
        soup1 = BeautifulSoup(page1.text, "html.parser")
        block1 = soup1.findAll('article', class_='product-card product-card product')

        for data in block1:
            name = data.find(class_='product-title__head')
            author = data.find(class_='product-title__author')
            price = data.find(class_='product-price__value')
            if (name and author and price) is not None:
                description1.append(name.text)
                description2.append(author.text)
                description3.append(price.text)

    print(description1, description2, description3)
    pozitionNow = 1
    for elem1, elem2, elem3 in zip(description1, description2, description3):
        cell = spisok.cell(1, pozitionNow)
        cell.value = elem1

        cell = spisok.cell(2, pozitionNow)
        cell.value = elem2

        cell = spisok.cell(3, pozitionNow)
        cell.value = elem3

        pozitionNow += 1from bs4 import BeautifulSoup  # Импорт модуля BeautifulSoup из библиотеки bs4

import requests  # Импорт модуля requests

from openpyxl import load_workbook  # Импорт функции load_workbook из модуля openpyxl

import os  # Импорт модуля os

def parse1():

    filename = 'лабалоторная1.xlsx'  # Имя файла Excel

    exelfile = load_workbook(filename)  # Загрузка файла Excel

    spisok = exelfile['данные']  # Получение доступа к листу данных 'данные' в Excel-файле

    description1 = []  # Пустой список для хранения данных (название)

    description2 = []  # Пустой список для хранения данных (автор)

    description3 = []  # Пустой список для хранения данных (цена)

    count = 2  # Количество страниц для парсинга

    ssilka = 'https://www.chitai-gorod.ru/catalog/collections/bestsell?page='  # Базовая ссылка для формирования URL-адресов страниц

    for i in range(1, count + 1):  # Цикл для парсинга каждой страницы

        url1 = ssilka + str(i)  # Формирование URL-адреса страницы

        page1 = requests.get(url1)  # Отправка GET-запроса к странице

        print(page1.status_code)  # Проверка статуса кода ответа

        soup1 = BeautifulSoup(page1.text, "html.parser")  # Создание объекта BeautifulSoup для парсинга HTML-кода страницы

        block1 = soup1.findAll('article', class_='product-card product-card product')  # Поиск всех элементов <article> с указанными классами

        for data in block1:  # Цикл для извлечения данных из каждого элемента

            name = data.find(class_='product-title__head')  # Поиск элемента с указанным классом (название)

            author = data.find(class_='product-title__author')  # Поиск элемента с указанным классом (автор)

            price = data.find(class_='product-price__value')  # Поиск элемента с указанным классом (цена)

            if (name and author and price) is not None:  # Проверка, что все данные присутствуют

                description1.append(name.text)  # Добавление названия в список

                description2.append(author.text)  # Добавление автора в список

                description3.append(price.text)  # Добавление цены в список

    print(description1, description2, description3)  # Вывод списков данных

    pozitionNow = 1  # Начальная позиция для записи в Excel

    for elem1, elem2, elem3 in zip(description1, description2, description3):  # Цикл для записи данных в Excel

        cell = spisok.cell(1, pozitionNow)

    exelfile.save(filename)
    exelfile        .close()

    # Открывает файл Excel после сохранения
    os.startfile(filename)


parse1()
